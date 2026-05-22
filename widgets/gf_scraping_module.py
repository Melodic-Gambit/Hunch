"""
Встроенный модуль GF.Scraping для вкладки «Сервисы» Hunch.exe.
"""
import sys
import hashlib
import tkinter as tk
from tkinter import filedialog
import customtkinter as ctk
from PIL import Image
import threading
from datetime import datetime
import os
import time
from typing import Optional, Callable
import theme_colors


def _teal():
    return (theme_colors.accent(), theme_colors.hover())


def _teal_hvr():
    return (theme_colors.hover(), theme_colors.dark())
_RED      = ("#EF4444", "#DC2626")
_RED_HVR  = ("#DC2626", "#B91C1C")
_GRAY_BTN = ("gray55", "gray35")
_GRAY_HVR = ("gray45", "gray25")


def _show_gf_warning(parent_widget: ctk.CTkBaseClass, message: str):
    """Стилизованный диалог предупреждения вместо стандартного messagebox."""
    parent = parent_widget.winfo_toplevel()
    dlg = ctk.CTkToplevel(parent)
    dlg.withdraw()
    dlg.title("Предупреждение")
    dlg.resizable(False, False)
    dlg.transient(parent)
    dlg.grab_set()
    dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)

    ctk.CTkLabel(
        dlg,
        text="⚠  " + message,
        font=ctk.CTkFont(size=13),
        text_color=("gray10", "white"),
        anchor="w",
        justify="left",
        wraplength=320,
    ).pack(padx=28, pady=(26, 14))

    ctk.CTkButton(
        dlg, text="OK", width=90, height=32,
        fg_color=_teal(), hover_color=_teal_hvr(),
        command=dlg.destroy,
    ).pack(pady=(0, 22))

    def _center():
        dlg.update_idletasks()
        parent.update_idletasks()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        if pw <= 1 or ph <= 1:
            dlg.after(80, _center)
            return
        dw = dlg.winfo_width()
        dh = dlg.winfo_height()
        x = parent.winfo_rootx() + (pw - dw) // 2
        y = parent.winfo_rooty() + (ph - dh) // 2
        dlg.geometry(f"+{x}+{y}")
        dlg.deiconify()

    dlg.after(60, _center)
    parent.wait_window(dlg)


_GF_SCRAPING_VERSION = "v1.0"
_GF_DEFAULT_TIMEOUT = 15
_GF_RETRY_DELAYS    = (1, 2, 4)  # секунды между попытками (exponential backoff)


def _gf_requests_get(url: str, timeout: int = _GF_DEFAULT_TIMEOUT,
                     headers: dict = None) -> "requests.Response":
    """GET с exponential-backoff retry (3 попытки)."""
    import requests as _req
    last_exc = None
    for attempt, delay in enumerate(_GF_RETRY_DELAYS):
        try:
            resp = _req.get(url, timeout=timeout, headers=headers or {})
            resp.raise_for_status()
            return resp
        except Exception as exc:
            last_exc = exc
            if attempt < len(_GF_RETRY_DELAYS) - 1:
                time.sleep(delay)
    raise last_exc


def _gf_check_url_hash(url: str, timeout: int = _GF_DEFAULT_TIMEOUT) -> Optional[str]:
    """Возвращает MD5-хэш таблицы ОКПД/ОКВЭД с URL или None при ошибке.
    Используется для фонового мониторинга изменений."""
    try:
        from bs4 import BeautifulSoup
        resp = _gf_requests_get(url, timeout=timeout)
        soup  = BeautifulSoup(resp.content, "html.parser")
        table = soup.find("table",
                          class_="table table-bordered table-width-auto")
        if table is None:
            return None
        return hashlib.md5(table.get_text(separator="|").encode()).hexdigest()
    except Exception:
        return None


def _gf_fetch_latest_numbers(page_type: str,
                              timeout: int = _GF_DEFAULT_TIMEOUT) -> list:
    """Возвращает список номеров из колонки «Номер» индексной страницы ОКВЭД/ОКПД.

    page_type: 'okved' или 'okpd'
    Возвращает list[int] — все номера из таблицы, сортировка по убыванию.
    """
    url = f"https://classifikators.ru/updates/{page_type}"
    try:
        from bs4 import BeautifulSoup
        resp = _gf_requests_get(url, timeout=timeout, headers={
            "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/124.0 Safari/537.36"),
            "Accept-Language": "ru-RU,ru;q=0.9",
        })
        soup = BeautifulSoup(resp.content, "html.parser")

        # Ищем таблицу: сначала по точному классу, затем по частичному
        table = soup.find("table", class_="table table-bordered table-width-auto")
        if not table:
            table = soup.find("table", class_="table-bordered")
        if not table:
            # Последний вариант: первая <table> на странице
            table = soup.find("table")
        if not table:
            return []

        tbody = table.find("tbody")
        if not tbody:
            return []

        # Определяем индекс колонки «Номер»
        num_col = 0
        thead = table.find("thead")
        if thead:
            headers = [th.get_text(strip=True) for th in thead.find_all("th")]
            if "Номер" in headers:
                num_col = headers.index("Номер")

        numbers = []
        for tr in tbody.find_all("tr"):
            cols = tr.find_all("td")
            if len(cols) > num_col:
                text = cols[num_col].get_text(strip=True)
                part = text.split("/")[0].strip()
                if part.isdigit():
                    numbers.append(int(part))
        return sorted(set(numbers), reverse=True)
    except Exception:
        return []


class GFScrapingFrame(ctk.CTkFrame):
    """UI GF.Scraping встроенный во фрейм — без собственного mainloop."""

    # Поля для форм ID: (label, settings_key)
    _ID_FIELDS = {
        "hotfix": [
            ("OKPD: Начальный id (hotfix)",                  "start_id_hotfix"),
            ("OKVED: Начальный id text (hotfix)",            "start_id_text_hotfix"),
            ("OKVED: Начальный id language_text (hotfix)",   "start_id_language_text_hotfix"),
            ("OKVED: Начальный id okved_types (hotfix)",     "start_id_okved_types_hotfix"),
        ],
        "app": [
            ("OKPD: Начальный id (app)",                     "start_id_app"),
            ("OKVED: Начальный id text (app)",               "start_id_text_app"),
            ("OKVED: Начальный id language_text (app)",      "start_id_language_text_app"),
            ("OKVED: Начальный id okved_types (app)",        "start_id_okved_types_app"),
        ],
    }

    def __init__(self, parent,
                 settings_manager=None,
                 log_manager:    object   = None,
                 notify_cb:      Callable = None,
                 version:        str      = "",
                 **kw):
        super().__init__(parent, fg_color="transparent", **kw)
        self._settings   = settings_manager
        self._log_mgr    = log_manager
        self._notify_cb  = notify_cb
        self._version    = version
        self._destroyed  = False
        self._stop_event = threading.Event()
        self.urls:       list = []
        self.entries:    dict = {}
        self.field_rows: list = []
        self.timeout_var = tk.StringVar(value=str(_GF_DEFAULT_TIMEOUT))
        self._create_widgets()
        self._restore_state()

    # ── state ──────────────────────────────────────────────────────────────────

    def _get_timeout(self) -> int:
        try:
            val = int(self.timeout_var.get())
            return max(5, min(60, val))
        except Exception:
            return _GF_DEFAULT_TIMEOUT

    def _restore_state(self):
        if self._settings is None:
            return
        d = self._settings.get_setting("gf_scraping_state", {})
        for url in d.get("urls", []):
            if url not in self.urls:
                self.urls.append(url)
                self.url_listbox.insert(tk.END, url)
        sp = d.get("save_path", "")
        if sp:
            self.save_path_var.set(sp)
        t = d.get("timeout", _GF_DEFAULT_TIMEOUT)
        self.timeout_var.set(str(t))

    def _save_state(self):
        if self._settings is None:
            return
        self._settings.set_setting("gf_scraping_state", {
            "urls":      list(self.urls),
            "save_path": self.save_path_var.get(),
            "timeout":   self._get_timeout(),
        })

    # ── thread-safe helpers ────────────────────────────────────────────────────

    def _log(self, message: str):
        """Отправляет сообщение в LogManager основного приложения (thread-safe)."""
        if self._destroyed:
            return
        level = ("ERROR"   if ("❌" in message or "🚨" in message
                                or "ошибка" in message.lower()) else
                 "WARNING" if ("⚠️" in message
                                or "предупреждение" in message.lower()) else
                 "INFO")
        if self._log_mgr is not None:
            try:
                self.after(0, lambda m=message, lv=level:
                           (None if self._destroyed
                            else self._log_mgr.add_log(f"[GF.Scraping] {m}", lv)))
            except Exception:
                pass

    def _ui(self, fn: Callable):
        """Выполнить fn в главном потоке."""
        if self._destroyed:
            return
        try:
            self.after(0, fn)
        except Exception:
            pass

    # ── tooltip ────────────────────────────────────────────────────────────────

    def _show_tooltip(self, event, text: str):
        top = tk.Toplevel(self)
        top.wm_overrideredirect(True)
        top.wm_geometry(f"+{event.x_root + 12}+{event.y_root + 12}")
        tk.Label(top, text=text,
                 background="#1A1A2E", foreground="white",
                 relief="flat", bd=0,
                 font=("Segoe UI", 10), wraplength=420,
                 justify="left", padx=10, pady=6).pack()
        event.widget._tooltip = top

    def _hide_tooltip(self, event):
        t = getattr(event.widget, "_tooltip", None)
        if t:
            try:
                t.destroy()
            except Exception:
                pass
            event.widget._tooltip = None

    # ── URL management ─────────────────────────────────────────────────────────

    def add_url(self):
        url = self.url_entry.get().strip()
        if not url:
            return
        if url in self.urls:
            self._log(f"⚠️ Ссылка уже в списке: {url}")
            return
        self.urls.append(url)
        self.url_listbox.insert(tk.END, url)
        self.url_entry.delete(0, tk.END)
        self._log(f"✅ Добавлена: {url}")
        self._save_state()

    def edit_url(self):
        sel = self.url_listbox.curselection()
        if not sel:
            return
        new_url = self.url_entry.get().strip()
        if not new_url:
            return
        idx = sel[0]
        old = self.urls[idx]
        self.urls[idx] = new_url
        self.url_listbox.delete(idx)
        self.url_listbox.insert(idx, new_url)
        self._log(f"✅ Изменена: {old} → {new_url}")
        self._save_state()

    def delete_url(self):
        sel = self.url_listbox.curselection()
        if not sel:
            return
        idx     = sel[0]
        deleted = self.urls.pop(idx)
        self.url_listbox.delete(idx)
        self._log(f"✅ Удалена: {deleted}")
        self._save_state()

    # ── widget creation ────────────────────────────────────────────────────────

    def _create_widgets(self):
        # ── header bar ────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=_teal(), corner_radius=0, height=50)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        _base = (sys._MEIPASS if getattr(sys, "frozen", False)
                 else os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        try:
            _pil  = Image.open(os.path.join(_base, "gf_logo.png"))
            _logo = ctk.CTkImage(light_image=_pil, dark_image=_pil, size=(26, 26))
            ctk.CTkLabel(hdr, image=_logo, text="",
                         fg_color="transparent").pack(side="left", padx=(14, 6))
        except Exception:
            ctk.CTkLabel(hdr, text="🔍", fg_color="transparent",
                         font=ctk.CTkFont(size=18), text_color="white"
                         ).pack(side="left", padx=(14, 6))

        ctk.CTkLabel(hdr, text="GF. Scraping — ОКПД / ОКВЭД",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     fg_color="transparent", text_color="white"
                     ).pack(side="left")
        ctk.CTkLabel(hdr, text=self._version or _GF_SCRAPING_VERSION,
                     font=ctk.CTkFont(size=11),
                     fg_color="transparent",
                     text_color=("white", "#CCF2EE")
                     ).pack(side="left", padx=(6, 0))

        # ── scrollable content ────────────────────────────────────────────────
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=14, pady=10)
        scroll.grid_columnconfigure(0, weight=1)

        row = 0

        # ── URL input ─────────────────────────────────────────────────────────
        url_sec = ctk.CTkFrame(scroll)
        url_sec.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        url_sec.grid_columnconfigure(0, weight=1)
        row += 1

        hdr_row = ctk.CTkFrame(url_sec, fg_color="transparent")
        hdr_row.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 0))
        ctk.CTkLabel(hdr_row, text="Ссылка на справочник",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     anchor="w").pack(side="left")
        _tip = ("Введите ссылку списка ОКПД/ОКВЭД с сайта classifikators.ru.\n"
                "Пример ОКПД:  https://classifikators.ru/updates/okpd/130\n"
                "Пример ОКВЭД: https://classifikators.ru/updates/okved/84")
        tip = ctk.CTkLabel(hdr_row, text="ℹ️", text_color="gray",
                           cursor="hand2", font=ctk.CTkFont(size=12))
        tip.pack(side="right")
        tip.bind("<Enter>", lambda e: self._show_tooltip(e, _tip))
        tip.bind("<Leave>", self._hide_tooltip)

        self.url_entry = ctk.CTkEntry(
            url_sec,
            placeholder_text="https://classifikators.ru/updates/okpd/...",
            height=34)
        self.url_entry.grid(row=1, column=0, sticky="ew", padx=12, pady=(5, 2))
        self.url_entry.bind("<Return>", lambda _: self.add_url())

        # Ctrl+V / Ctrl+М (рус) — вставка независимо от раскладки
        def _paste_url(event=None):
            try:
                text = self.clipboard_get()
            except Exception:
                return "break"
            if text:
                self.url_entry.delete(0, tk.END)
                self.url_entry.insert(0, text)
            return "break"
        self.url_entry.bind("<Control-v>", _paste_url)
        self.url_entry.bind("<Control-V>", _paste_url)
        self.url_entry.bind("<Control-KeyPress>",
            lambda e: _paste_url() if (e.keycode == 86
                                        and e.keysym.lower() not in ("v",))
                      else None)

        ctx = tk.Menu(self, tearoff=0)
        ctx.add_command(label="Вставить", command=_paste_url)
        self.url_entry.bind("<Button-3>",
                            lambda e: ctx.tk_popup(e.x_root, e.y_root))

        ctk.CTkLabel(url_sec,
                     text="↵ Enter — добавить  •  ПКМ — вставить",
                     text_color="gray", font=ctk.CTkFont(size=10), anchor="w"
                     ).grid(row=2, column=0, sticky="w", padx=14, pady=(0, 4))

        btn_r = ctk.CTkFrame(url_sec, fg_color="transparent")
        btn_r.grid(row=3, column=0, sticky="w", padx=12, pady=(0, 10))
        ctk.CTkButton(btn_r, text="Добавить", command=self.add_url,
                      width=88, height=28,
                      fg_color=_teal(), hover_color=_teal_hvr()
                      ).pack(side="left", padx=(0, 6))

        # ── URL list ──────────────────────────────────────────────────────────
        list_sec = ctk.CTkFrame(scroll)
        list_sec.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        row += 1

        ctk.CTkLabel(list_sec, text="Список ссылок",
                     font=ctk.CTkFont(size=12, weight="bold"), anchor="w"
                     ).pack(anchor="w", padx=12, pady=(8, 4))
        self.url_listbox = tk.Listbox(
            list_sec, height=4,
            bg="#2b2b2b", fg="white",
            selectbackground=theme_colors.accent(), selectforeground="white",
            activestyle="none", relief="flat", bd=0,
            font=("Segoe UI", 11))
        self.url_listbox.pack(fill="x", padx=12, pady=(0, 4))
        ctk.CTkButton(list_sec, text="Удалить выбранную",
                      command=self.delete_url,
                      width=150, height=26,
                      fg_color=_RED, hover_color=_RED_HVR
                      ).pack(anchor="w", padx=12, pady=(0, 10))

        # ── checkboxes ────────────────────────────────────────────────────────
        cb_sec = ctk.CTkFrame(scroll)
        cb_sec.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        row += 1

        ctk.CTkLabel(cb_sec, text="Тип идентификаторов",
                     font=ctk.CTkFont(size=12, weight="bold"), anchor="w"
                     ).pack(anchor="w", padx=12, pady=(8, 4))

        self.id_app_var    = tk.BooleanVar(value=True)
        self.id_hotfix_var = tk.BooleanVar(value=True)
        self.okpd_var      = tk.BooleanVar(value=True)
        self.okved_var     = tk.BooleanVar(value=True)

        cb_r = ctk.CTkFrame(cb_sec, fg_color="transparent")
        cb_r.pack(fill="x", padx=12, pady=(0, 10))
        for text, var in (("APP",    self.id_app_var),
                          ("HOTFIX", self.id_hotfix_var),
                          ("OKPD",   self.okpd_var),
                          ("OKVED",  self.okved_var)):
            ctk.CTkCheckBox(cb_r, text=text, variable=var,
                            command=self.toggle_id_fields,
                            fg_color=theme_colors.accent(), hover_color=theme_colors.hover(),
                            checkmark_color="white"
                            ).pack(side="left", padx=(0, 12))

        # По умолчанию показываем все поля (все чекбоксы включены)
        self.after(0, self.toggle_id_fields)

        # ── ID fields (scrollable) ────────────────────────────────────────────
        id_sec = ctk.CTkFrame(scroll)
        id_sec.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        id_sec.grid_columnconfigure(0, weight=1)
        row += 1

        self.form_frame = ctk.CTkScrollableFrame(id_sec, height=170,
                                                  fg_color="transparent")
        self.form_frame.pack(fill="x", padx=12, pady=8)
        self.form_frame.grid_columnconfigure(0, weight=1)

        # ── save path ─────────────────────────────────────────────────────────
        save_sec = ctk.CTkFrame(scroll)
        save_sec.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        save_sec.grid_columnconfigure(1, weight=1)
        row += 1

        ctk.CTkLabel(save_sec, text="Сохранить в:",
                     font=ctk.CTkFont(size=12, weight="bold")
                     ).grid(row=0, column=0, sticky="w",
                            padx=(12, 8), pady=10)
        self.save_path_var = tk.StringVar()
        ctk.CTkEntry(save_sec, textvariable=self.save_path_var,
                     placeholder_text="Папка не выбрана",
                     state="readonly", height=30
                     ).grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=10)
        ctk.CTkButton(save_sec, text="Обзор",
                      command=self._browse_save_directory,
                      width=70, height=30,
                      fg_color=_GRAY_BTN, hover_color=_GRAY_HVR
                      ).grid(row=0, column=2, sticky="e", padx=(0, 12), pady=10)

        # ── network settings ──────────────────────────────────────────────────
        net_sec = ctk.CTkFrame(scroll)
        net_sec.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        net_sec.grid_columnconfigure(1, weight=1)
        row += 1

        ctk.CTkLabel(net_sec, text="Настройки сети",
                     font=ctk.CTkFont(size=12, weight="bold"), anchor="w"
                     ).grid(row=0, column=0, columnspan=3, sticky="w",
                            padx=12, pady=(8, 4))

        ctk.CTkLabel(net_sec, text="Таймаут (сек, 5–60):", anchor="w"
                     ).grid(row=1, column=0, sticky="w", padx=(12, 8), pady=(0, 10))
        ctk.CTkEntry(net_sec, textvariable=self.timeout_var, width=70, height=28
                     ).grid(row=1, column=1, sticky="w", pady=(0, 10))
        ctk.CTkLabel(net_sec,
                     text="Авто-повтор: 3 попытки (1 с → 2 с → 4 с)",
                     text_color="gray", font=ctk.CTkFont(size=10), anchor="w"
                     ).grid(row=1, column=2, sticky="w", padx=(12, 12), pady=(0, 10))

        # ── progress ──────────────────────────────────────────────────────────
        prog_sec = ctk.CTkFrame(scroll, fg_color="transparent")
        prog_sec.grid(row=row, column=0, sticky="ew", pady=(0, 4))
        row += 1

        self.progress = ctk.CTkProgressBar(prog_sec, mode="determinate",
                                            progress_color=theme_colors.accent())
        self.progress.pack(fill="x", padx=14, pady=(0, 3))
        self.progress.set(0)
        self.progress_label = ctk.CTkLabel(
            prog_sec, text="Ожидает",
            font=ctk.CTkFont(size=11),
            text_color=("gray40", "gray70"), anchor="w")
        self.progress_label.pack(fill="x", padx=14)

        # ── launch / stop buttons ─────────────────────────────────────────────
        btn_row = ctk.CTkFrame(scroll, fg_color="transparent", height=1)
        btn_row.grid(row=row, column=0, sticky="ew", padx=14, pady=(4, 8))
        btn_row.grid_columnconfigure(0, weight=1)

        self.run_button = ctk.CTkButton(
            btn_row, text="🚀  Запустить парсинг",
            command=self.start_parsing,
            height=40, font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=_teal(), hover_color=_teal_hvr())
        self.run_button.grid(row=0, column=0, sticky="ew", padx=(0, 6))

        self.stop_button = ctk.CTkButton(
            btn_row, text="■  Остановить",
            command=self._stop_parsing,
            height=40, width=130, font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=_RED, hover_color=_RED_HVR,
            state="disabled")
        self.stop_button.grid(row=0, column=1, sticky="e")

    # ── id fields ──────────────────────────────────────────────────────────────

    def toggle_id_fields(self):
        # Уничтожить старые строки (entries — дочерние элементы, удаляются вместе с row)
        for row_w in self.field_rows:
            try:
                row_w.destroy()
            except Exception:
                pass
        self.field_rows.clear()
        self.entries.clear()

        # Определяем фильтры по mode (APP/HOTFIX) и type (OKPD/OKVED)
        modes = [m for m, v in (("app",    self.id_app_var),
                                 ("hotfix", self.id_hotfix_var)) if v.get()]
        types = [t for t, v in (("OKPD",   self.okpd_var),
                                 ("OKVED",  self.okved_var)) if v.get()]

        # Ничего не выбрано — ничего не показываем
        if not modes and not types:
            return

        # Нет фильтра по mode — показываем оба; нет фильтра по type — показываем оба
        show_modes = modes if modes else ["app", "hotfix"]
        show_types = types if types else ["OKPD", "OKVED"]

        for mode in show_modes:
            for ft in show_types:
                for label_text, key in self._ID_FIELDS[mode]:
                    if f"{ft}:" not in label_text or key in self.entries:
                        continue
                    row = ctk.CTkFrame(self.form_frame, fg_color="transparent")
                    row.pack(fill="x", pady=2, padx=4)
                    row.grid_columnconfigure(0, weight=1)
                    ctk.CTkLabel(row, text=label_text,
                                 anchor="w", font=ctk.CTkFont(size=11)
                                 ).grid(row=0, column=0, sticky="w")
                    entry = ctk.CTkEntry(row, placeholder_text="0",
                                         width=110, height=26)
                    entry.grid(row=0, column=1, sticky="e", padx=(8, 0))
                    self.entries[key] = entry
                    self.field_rows.append(row)

    # ── parsing ────────────────────────────────────────────────────────────────

    def start_parsing(self):
        try:
            ids = {k: int(e.get())
                   for k, e in self.entries.items() if e.get().strip()}
        except ValueError:
            from tkinter import messagebox as _mb
            _mb.showerror("Ошибка",
                          "Все заполненные поля должны содержать целые числа!")
            return
        if not self.urls:
            from tkinter import messagebox as _mb
            _mb.showerror("Ошибка", "Добавьте хотя бы один URL!")
            return
        save_dir = self.save_path_var.get()
        if not save_dir:
            _show_gf_warning(self, "Выберите директорию для сохранения файла.")
            return

        # Захватываем данные до запуска потока — избегаем чтения UI из потока
        urls_copy     = list(self.urls)
        stored_hashes = {}
        if self._settings is not None:
            stored_hashes = dict(
                self._settings.get_setting("gf_scraping_hashes", {}))

        self._stop_event.clear()
        self.run_button.configure(state="disabled")
        self.stop_button.configure(state="normal", text="■  Остановить")
        self.progress.set(0)
        self.progress.configure(progress_color=theme_colors.accent())
        self.progress_label.configure(text=f"Парсинг 0/{len(urls_copy)}")
        threading.Thread(
            target=self._run_scraping,
            args=(ids, urls_copy, save_dir, stored_hashes, self._get_timeout()),
            daemon=True,
        ).start()

    def _stop_parsing(self):
        self._stop_event.set()
        try:
            self.stop_button.configure(state="disabled", text="Остановка…")
        except Exception:
            pass

    def _browse_save_directory(self):
        d = filedialog.askdirectory(title="Выберите директорию для сохранения")
        if d:
            self.save_path_var.set(d)
            self._log(f"✅ Директория: {d}")
            self._save_state()
        else:
            self._log("⚠️ Директория не выбрана")

    def _run_scraping(self, ids: dict, urls: list,
                      save_dir: str, stored_hashes: dict, net_timeout: int = _GF_DEFAULT_TIMEOUT):
        try:
            from bs4 import BeautifulSoup
            import pandas as pd
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError as _imp_err:
            self._log(f"🚨 Ошибка: не установлен пакет — {_imp_err}")
            self._log(
                "Установите: pip install requests beautifulsoup4 pandas openpyxl")
            self._ui(lambda: (
                self.run_button.configure(state="normal"),
                self.progress_label.configure(text="Ошибка зависимостей"),
            ))
            return

        try:
            cur = {
                "id_hotfix":             ids.get("start_id_hotfix", 1),
                "id_app":                ids.get("start_id_app", 1),
                "id_text_hotfix":        ids.get("start_id_text_hotfix", 1),
                "id_text_app":           ids.get("start_id_text_app", 1),
                "id_lang_hotfix":        ids.get("start_id_language_text_hotfix", 1),
                "id_lang_app":           ids.get("start_id_language_text_app", 1),
                "id_okved_types_hotfix": ids.get("start_id_okved_types_hotfix", 1),
                "id_okved_types_app":    ids.get("start_id_okved_types_app", 1),
            }

            output_path = os.path.join(
                save_dir,
                f"{datetime.now().strftime('%d.%m.%Y')}.output_data.xlsx")
            os.makedirs(save_dir, exist_ok=True)
            self._log(f"Файл будет сохранён: {output_path}")
            total = len(urls)

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for i, url in enumerate(urls):
                    if self._stop_event.is_set():
                        break
                    self._log(f"Обработка ({i + 1}/{total}): {url}")
                    try:
                        resp = _gf_requests_get(url, timeout=net_timeout)
                        soup  = BeautifulSoup(resp.content, "html.parser")
                        table = soup.find(
                            "table",
                            class_="table table-bordered table-width-auto")
                        if not table:
                            self._log(f"❌ Таблица не найдена: {url}")
                            continue

                        thead   = table.find("thead")
                        headers = ([th.get_text(strip=True)
                                    for th in thead.find_all("th")[:3]]
                                   if thead else ["Кол1", "Кол2", "Кол3"])

                        data  = []
                        tbody = table.find("tbody")
                        if tbody:
                            for tr in tbody.find_all("tr"):
                                cols = tr.find_all("td")
                                if len(cols) >= 3:
                                    data.append([
                                        cols[j].get_text(strip=True)
                                               .split("детали ↓")[0]
                                        for j in range(3)
                                    ])

                        # ── уведомление об изменениях ─────────────────────────
                        if self._notify_cb and data:
                            new_hash = hashlib.md5(
                                str(data).encode()).hexdigest()
                            old_hash = stored_hashes.get(url)
                            self._ui(lambda u=url, oh=old_hash,
                                            nh=new_hash, c=len(data):
                                     self._notify_cb(u, oh, nh, c))
                            stored_hashes[url] = new_hash

                        df = pd.DataFrame(data, columns=headers)
                        df["SQL hotfix.goodfin"] = ""
                        df["SQL app.goodfin"]    = ""

                        parts      = url.rstrip("/").split("/")[-2:]
                        sheet_name = "-".join(parts)[:31]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        ws      = writer.sheets[sheet_name]
                        col_hf  = len(headers) + 1
                        col_app = len(headers) + 2
                        ws.column_dimensions[get_column_letter(col_hf)].width  = 100
                        ws.column_dimensions[get_column_letter(col_app)].width = 100

                        if not data:
                            self._log(f"⚠️ Нет данных для {sheet_name}")
                            continue

                        r_start = 2
                        r_end   = len(data) + 1
                        align   = openpyxl.styles.Alignment(
                            wrap_text=True, vertical="top")

                        if "okpd" in url.lower():
                            lines_hf, lines_app = [], []
                            for row in data:
                                if row[0] == "В" and row[1] and row[2]:
                                    code = row[1].strip()
                                    name = row[2].strip()
                                    lines_hf.append(
                                        f"({cur['id_hotfix']}, ???, "
                                        f"'{code}', '1', '{name}')")
                                    lines_app.append(
                                        f"({cur['id_app']}, ???, "
                                        f"'{code}', '1', '{name}')")
                                    cur["id_hotfix"] += 1
                                    cur["id_app"]    += 1
                            for lines, col in ((lines_hf,  col_hf),
                                               (lines_app, col_app)):
                                if lines:
                                    c_ref = f"{get_column_letter(col)}{r_start}"
                                    ws.merge_cells(
                                        f"{c_ref}:"
                                        f"{get_column_letter(col)}{r_end}")
                                    cell = ws[c_ref]
                                    cell.value = (
                                        "INSERT INTO public.okpd_types\n"
                                        "(id, parentid, code, state, name)"
                                        "\nVALUES\n"
                                        + ",\n".join(lines))
                                    cell.alignment = align

                        elif "okved" in url.lower():
                            blocks_hf, blocks_app = [], []
                            for row in data:
                                if row[0] == "В" and row[1] and row[2]:
                                    code = row[1].strip()
                                    name = row[2].strip()
                                    sc   = code.replace(".", "")
                                    for blist, id_text, id_lang, id_types in (
                                        (blocks_hf,
                                         "id_text_hotfix", "id_lang_hotfix",
                                         "id_okved_types_hotfix"),
                                        (blocks_app,
                                         "id_text_app", "id_lang_app",
                                         "id_okved_types_app"),
                                    ):
                                        blist.append(
                                            f"INSERT INTO public.\"text\" "
                                            f"(id) VALUES ({cur[id_text]}); "
                                            f"INSERT INTO public.language_text "
                                            f"(id, textid, languagetypeid,"
                                            f" localizedtext) VALUES "
                                            f"({cur[id_lang]}, {cur[id_text]},"
                                            f" 'ru', '{name}'); "
                                            f"INSERT INTO public.okved_types "
                                            f"(id, parentid, code, state,"
                                            f" searchcode, nametextid) VALUES "
                                            f"({cur[id_types]}, ???, '{code}',"
                                            f" '1', '{sc}', {cur[id_text]})"
                                        )
                                        cur[id_text]  += 1
                                        cur[id_lang]  += 1
                                        cur[id_types] += 1
                            for blocks, col in ((blocks_hf,  col_hf),
                                                (blocks_app, col_app)):
                                if blocks:
                                    c_ref = f"{get_column_letter(col)}{r_start}"
                                    ws.merge_cells(
                                        f"{c_ref}:"
                                        f"{get_column_letter(col)}{r_end}")
                                    cell = ws[c_ref]
                                    cell.value = "\n\n".join(
                                        s.strip() + ";"
                                        for s in ";".join(blocks).split(";")
                                        if s.strip())
                                    cell.alignment = align

                        n = i + 1
                        self._ui(lambda n=n, t=total: (
                            self.progress.set(n / t),
                            self.progress_label.configure(
                                text=f"Обработано {n}/{t}"),
                        ))
                        self._log(f"✅ Лист '{sheet_name}' сохранён")

                        if i < total - 1:
                            if self._stop_event.wait(11):
                                break

                    except Exception as exc:
                        self._log(f"❌ Ошибка при обработке {url}: {exc}")

            self._log(f"🎉 Готово! Файл: {output_path}")

        except Exception as exc:
            self._log(f"🚨 Критическая ошибка: {exc}")

        finally:
            stopped = self._stop_event.is_set()
            self._ui(lambda s=stopped: (
                self.progress.set(1.0),
                self.progress.configure(
                    progress_color="#EF4444" if s else "#22C55E"),
                self.progress_label.configure(
                    text="Остановлено" if s else "Готово"),
                self.run_button.configure(state="normal"),
                self.stop_button.configure(
                    state="disabled", text="■  Остановить"),
            ))


# ── window wrapper ─────────────────────────────────────────────────────────────

class GFScrapingWindow(ctk.CTkToplevel):
    """Модальное окно GF.Scraping, открываемое из Hunch.exe."""

    _instance = None

    @classmethod
    def open(cls, parent,
             settings_manager=None,
             log_manager:     object   = None,
             notify_cb:       Callable = None,
             version:         str      = ""):
        """Открыть или поднять окно GF.Scraping."""
        if cls._instance is not None:
            try:
                cls._instance.deiconify()
                cls._instance.lift()
                cls._instance.focus_force()
                return cls._instance
            except Exception:
                cls._instance = None
        win = cls(parent,
                  settings_manager=settings_manager,
                  log_manager=log_manager,
                  notify_cb=notify_cb,
                  version=version)
        cls._instance = win
        return win

    def __init__(self, parent,
                 settings_manager=None,
                 log_manager:     object   = None,
                 notify_cb:       Callable = None,
                 version:         str      = ""):
        super().__init__(parent)
        self.withdraw()

        self.title("GF. Scraping")
        self.geometry("720x650")
        self.minsize(600, 500)
        self.resizable(True, True)
        self.transient(parent)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._frame = GFScrapingFrame(
            self,
            settings_manager=settings_manager,
            log_manager=log_manager,
            notify_cb=notify_cb,
            version=version,
        )
        self._frame.pack(fill="both", expand=True)

        self.update_idletasks()
        self._place_center(parent)
        self.deiconify()
        self.grab_set()
        self.after(20, self.lift)
        # Устанавливаем иконку приложения (убирает синюю иконку по умолчанию)
        _ico = os.path.join(
            sys._MEIPASS if getattr(sys, "frozen", False)
            else os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "Hunch.ico"
        )
        self.after(250, lambda i=_ico: self.iconbitmap(i) if os.path.exists(i) else None)

    def _place_center(self, parent):
        w = max(self.winfo_reqwidth(),  720)
        h = max(self.winfo_reqheight(), 650)
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        x  = px + (pw - w) // 2
        y  = py + (ph - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _on_close(self):
        if self._frame:
            self._frame._destroyed = True
        GFScrapingWindow._instance = None
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()
