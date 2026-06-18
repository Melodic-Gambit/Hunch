"""
Логика выгрузки SQL → XLSX для сервиса «SQL Выгрузка».
Без UI и без планировщика: только фоновый поток выполнения запросов и записи файла.
"""
import os
import threading
from datetime import datetime
from typing import List, Dict, Callable, Optional

SQL_EXPORT_VERSION = "v1.0.0"


class SqlExportService:
    """Выполняет набор SQL-запросов и сохраняет результаты в XLSX-файл.

    Использует db_manager._run() напрямую — без whitelist-валидации,
    так как отчётные запросы могут использовать CTE, CALL и другие конструкции.
    """

    def __init__(self, db_manager, log_cb: Optional[Callable] = None):
        """
        db_manager — DatabaseManager приложения
        log_cb     — callable(msg: str, level: str) для записи в LogManager; может быть None
        """
        self._db     = db_manager
        self._log_cb = log_cb
        self._lock   = threading.Lock()
        self._running = False

    # ── публичный интерфейс ────────────────────────────────────────────────────

    def is_running(self) -> bool:
        with self._lock:
            return self._running

    def start(
        self,
        queries:      List[Dict],  # [{display_name, filepath, connections, enabled}]
        folder:       str,
        filename_tpl: str,
        file_mode:    str,         # "daily" | "cumulative"
        sheet_mode:   str,         # "per_query" | "single" | "aggregate"
        on_done:      Callable,    # on_done(filename: str, error: str | None)
    ) -> bool:
        """Запускает выгрузку в daemon-потоке. Возвращает False если уже выполняется."""
        with self._lock:
            if self._running:
                return False
            self._running = True
        threading.Thread(
            target=self._run,
            args=(queries, folder, filename_tpl, file_mode, sheet_mode, on_done),
            daemon=True,
        ).start()
        return True

    # ── внутренняя логика ──────────────────────────────────────────────────────

    def _log(self, msg: str, level: str = "INFO"):
        if self._log_cb:
            try:
                self._log_cb(f"[SQL Выгрузка] {msg}", level)
            except Exception:
                pass

    def _run(self, queries, folder, filename_tpl, file_mode, sheet_mode, on_done):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as e:
            with self._lock:
                self._running = False
            on_done(filename="", error=f"Не установлен openpyxl: {e}")
            return

        active = [q for q in queries if q.get("enabled", True)]
        if not active:
            with self._lock:
                self._running = False
            on_done(filename="", error="Нет активных запросов")
            return

        now  = datetime.now()
        base = filename_tpl.strip() or "отчёт"

        # ── имя файла зависит от режима ────────────────────────────────────────
        if file_mode == "cumulative":
            filename = f"{base}.xlsx"
        else:  # daily
            filename = f"{now.strftime('%d.%m.%Y.')} {base}.xlsx"

        # ── создаём папку ──────────────────────────────────────────────────────
        try:
            os.makedirs(folder, exist_ok=True)
        except Exception as e:
            with self._lock:
                self._running = False
            on_done(filename=filename, error=f"Путь недоступен: {folder} — {e}")
            return

        out_path = os.path.join(folder, filename)

        # ── открываем/создаём workbook ─────────────────────────────────────────
        if file_mode == "cumulative" and os.path.exists(out_path):
            try:
                wb = openpyxl.load_workbook(out_path)
            except Exception:
                wb = openpyxl.Workbook()
                _del_default_sheet(wb)
        else:
            wb = openpyxl.Workbook()
            _del_default_sheet(wb)

        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="0D9488")
        wrap_aln  = Alignment(wrap_text=True, vertical="top")
        date_str  = now.strftime("%d.%m.%Y")   # для имён листов и разделителей
        date_pfx  = now.strftime("%d.%m")       # короткий префикс для имён листов

        # ── single-sheet: берём существующий лист или создаём новый ───────────
        single_ws  = None
        single_row = 1
        if sheet_mode in ("single", "aggregate"):
            sn = "Выгрузка"
            if file_mode == "cumulative" and sn in wb.sheetnames:
                # дописываем в конец существующего листа
                single_ws  = wb[sn]
                single_row = (single_ws.max_row or 0) + 2
                # разделитель с датой
                cell = single_ws.cell(row=single_row, column=1,
                                      value=f"═══ {date_str} ═══")
                cell.font = Font(bold=True, size=11)
                single_row += 1
            else:
                if sn in wb.sheetnames:
                    del wb[sn]
                single_ws = wb.create_sheet(sn)

        # ── основной цикл по запросам ──────────────────────────────────────────
        for q in active:
            dname = q.get("display_name", "Query")
            fpath = q.get("filepath", "")

            # backward compat: старый формат "connection": str
            raw_conns = q.get("connections") or (
                [q.get("connection", "")] if q.get("connection") else []
            )
            conns = [c[:-5] if c.endswith(".json") else c for c in raw_conns if c]

            if not conns:
                self._log(f"{dname}: не задано ни одно подключение", "ERROR")
                if sheet_mode == "per_query":
                    _write_error_sheet(wb, dname, "Не задано подключение")
                else:
                    single_row = _write_error_block(
                        single_ws, single_row, dname, "Не задано подключение")
                continue

            self._log(f"Запрос: {dname} [{', '.join(conns)}]")

            # читаем SQL-файл
            try:
                with open(fpath, "r", encoding="utf-8-sig") as fh:
                    sql = fh.read().strip()
            except Exception as e:
                self._log(f"Ошибка чтения файла {fpath}: {e}", "ERROR")
                if sheet_mode == "per_query":
                    _write_error_sheet(wb, dname, str(e))
                else:
                    single_row = _write_error_block(
                        single_ws, single_row, dname, str(e))
                continue

            # выполняем запрос против каждого подключения
            all_results = []
            for conn in conns:
                try:
                    rows, cols = self._db._run(conn, sql)
                    all_results.append((rows, cols))
                    self._log(f"{dname} [{conn}]: {len(rows)} строк")
                except Exception as e:
                    self._log(f"Ошибка {dname} [{conn}]: {e}", "ERROR")

            if not all_results:
                if sheet_mode == "per_query":
                    _write_error_sheet(wb, dname, "Все подключения вернули ошибку")
                else:
                    single_row = _write_error_block(
                        single_ws, single_row, dname, "Все подключения вернули ошибку")
                continue

            # агрегация если режим "aggregate" и несколько результатов
            if sheet_mode == "aggregate" and len(all_results) > 1:
                rows, cols = _aggregate_results(all_results)
                self._log(f"{dname}: агрегировано {len(all_results)} подключений")
            else:
                rows, cols = all_results[0]

            # записываем результат
            if sheet_mode == "per_query":
                if file_mode == "cumulative":
                    sn = f"{date_pfx} {dname}"[:31]
                    counter = 1
                    base_sn = sn
                    while sn in wb.sheetnames:
                        sn = f"{base_sn[:28]}_{counter}"
                        counter += 1
                else:
                    sn = dname[:31]
                    if sn in wb.sheetnames:
                        del wb[sn]
                ws = wb.create_sheet(sn)
                _write_sheet(ws, cols, rows, hdr_font, hdr_fill, wrap_aln)
            else:
                single_row = _write_data_block(
                    single_ws, single_row, dname,
                    cols, rows, hdr_font, hdr_fill, wrap_aln)

        # гарантируем хотя бы один лист
        if not wb.sheetnames:
            wb.create_sheet("Выгрузка")

        # ── сохраняем файл ─────────────────────────────────────────────────────
        save_path = out_path
        try:
            wb.save(save_path)
        except PermissionError:
            # файл открыт в другой программе — сохраняем с суффиксом _1
            base, ext = os.path.splitext(save_path)
            save_path = f"{base}_1{ext}"
            try:
                wb.save(save_path)
                filename = os.path.basename(save_path)
            except Exception as e:
                with self._lock:
                    self._running = False
                on_done(filename=filename,
                        error=f"Не удалось сохранить файл: {e}")
                return
        except Exception as e:
            with self._lock:
                self._running = False
            on_done(filename=filename, error=f"Не удалось сохранить файл: {e}")
            return

        self._log(f"Файл сохранён: {save_path}")
        with self._lock:
            self._running = False
        on_done(filename=os.path.basename(save_path), error=None)


# ── helpers ────────────────────────────────────────────────────────────────────

def _del_default_sheet(wb):
    for name in ("Sheet", "Sheet1", "Лист1"):
        if name in wb.sheetnames:
            del wb[name]


def _aggregate_results(results):
    """Суммирует числовые ячейки из N одноструктурных результатов.

    Текстовые ячейки (не парсятся как float) берутся из первого результата.
    """
    base_rows, cols = results[0]
    if not base_rows:
        return base_rows, cols

    n_rows = len(base_rows)
    n_cols = len(base_rows[0])
    out = []

    for ri in range(n_rows):
        merged = []
        for ci in range(n_cols):
            base_val = base_rows[ri][ci]
            total = None
            for res_rows, _ in results:
                if ri >= len(res_rows):
                    total = None
                    break
                v = res_rows[ri][ci]
                try:
                    num = float(str(v).replace(" ", "").replace(",", "."))
                    total = (total or 0) + num
                except (ValueError, TypeError, AttributeError):
                    total = None
                    break
            if total is not None:
                try:
                    merged.append(int(total) if total == int(total) else total)
                except (OverflowError, ValueError):
                    merged.append(total)
            else:
                merged.append(base_val)
        out.append(tuple(merged))

    return out, cols


def _write_sheet(ws, cols, rows, hdr_font, hdr_fill, wrap_aln):
    """Записывает заголовки + данные на лист."""
    if cols:
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=str(c))
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = wrap_aln
    for ri, row in enumerate(rows, 2 if cols else 1):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)


def _write_data_block(ws, start_row, title, cols, rows, hdr_font, hdr_fill, wrap_aln) -> int:
    """Записывает блок (заголовок группы + данные) в single-лист. Возвращает следующую строку."""
    from openpyxl.styles import Font as _Font
    cell = ws.cell(row=start_row, column=1, value=f"=== {title} ===")
    cell.font = _Font(bold=True)
    start_row += 1
    if cols:
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=start_row, column=ci, value=str(c))
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = wrap_aln
        start_row += 1
    for row in rows:
        for ci, v in enumerate(row, 1):
            ws.cell(row=start_row, column=ci, value=v)
        start_row += 1
    return start_row + 1  # пустая строка-разделитель


def _write_error_sheet(wb, display_name: str, error_msg: str):
    """Создаёт лист с описанием ошибки."""
    sn = f"ОШИБКА — {display_name}"[:31]
    if sn in wb.sheetnames:
        del wb[sn]
    ws = wb.create_sheet(sn)
    ws.cell(row=1, column=1, value="Ошибка выполнения запроса")
    ws.cell(row=2, column=1, value=error_msg)


def _write_error_block(ws, start_row: int, display_name: str, error_msg: str) -> int:
    """Вставляет ошибку в single-лист. Возвращает следующую строку."""
    ws.cell(row=start_row, column=1, value=f"=== {display_name} — ОШИБКА ===")
    ws.cell(row=start_row + 1, column=1, value=error_msg)
    return start_row + 3
