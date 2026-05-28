import csv
import tkinter as tk
from tkinter import ttk
import customtkinter as ctk
from typing import Optional
import theme_colors

_PAGE_SIZE = 100


class ResultTable(ctk.CTkFrame):
    """Таблица результатов SQL: заголовки из cursor.description, сортировка, копирование, pagination."""

    _last_applied_mode: str = ""
    _last_applied_accent: str = ""

    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self._columns: list = []
        self._rows:    list = []
        self._sort_col: Optional[int] = None
        self._sort_rev: bool = False
        self._current_page: int = 0
        self._focused_col: Optional[str] = None
        self._focused_item: Optional[str] = None
        self._selected_cell_value: str = ""
        self._hidden_keys: set = set()
        self._hidden_rows: dict = {}
        self._hovered_item: str = ""
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self._build()

    def _build(self):
        ResultTable._apply_style()
        self._tree = ttk.Treeview(self, style="SS.Treeview",
                                   show="headings", selectmode="browse")
        self._vsb = ctk.CTkScrollbar(self, command=self._tree.yview)
        self._hsb = ctk.CTkScrollbar(self, orientation="horizontal",
                                      command=self._tree.xview)
        self._tree.configure(yscrollcommand=self._on_vsb_set,
                             xscrollcommand=self._on_hsb_set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        self._tree.bind("<Button-3>",      self._on_right_click)
        self._tree.bind("<Button-1>",      self._on_copy_click)
        self._tree.bind("<Button-1>",      self._on_cell_click, add="+")
        self._tree.bind("<Motion>",        self._on_motion)
        self._tree.bind("<Leave>",         self._on_leave_tree)
        self._tree.bind("<Control-c>",     self._copy_row)
        self._tree.bind("<Control-C>",     self._copy_row)
        self._tree.bind("<Control-KeyPress>", self._on_ctrl_keypress)
        self._tree.bind("<Prior>", lambda e: self._page_prev()  or "break")
        self._tree.bind("<Next>",  lambda e: self._page_next()  or "break")
        self._tree.bind("<Home>",  lambda e: self._page_first() or "break")
        self._tree.bind("<End>",   lambda e: self._page_last()  or "break")
        self._tree.bind("<MouseWheel>",       self._on_scroll_redraw, add="+")
        self._tree.bind("<<TreeviewScroll>>", self._on_scroll_redraw, add="+")

        # ── pagination bar (row=3, скрыта по умолчанию) ──────────────────────
        self._pag_bar = ctk.CTkFrame(self, fg_color="transparent", height=1)
        self._pag_bar.grid(row=3, column=0, columnspan=2, sticky="ew", padx=4, pady=(2, 2))
        self._pag_bar.grid_columnconfigure(1, weight=1)

        self._btn_prev = ctk.CTkButton(
            self._pag_bar, text="←", width=32, height=24,
            fg_color="transparent",
            hover_color=("gray70", "gray40"),
            text_color=("gray10", "gray90"),
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._page_prev,
        )
        self._btn_prev.grid(row=0, column=0, padx=(0, 4))

        self._pag_label = ctk.CTkLabel(
            self._pag_bar, text="1 / 1",
            font=ctk.CTkFont(size=12),
            text_color=("gray30", "gray70"),
            anchor="center",
        )
        self._pag_label.grid(row=0, column=1)

        self._btn_next = ctk.CTkButton(
            self._pag_bar, text="→", width=32, height=24,
            fg_color="transparent",
            hover_color=("gray70", "gray40"),
            text_color=("gray10", "gray90"),
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._page_next,
        )
        self._btn_next.grid(row=0, column=2, padx=(4, 0))

        self._pag_bar.grid_remove()

        # ── cell info bar (row=2, hidden by default) ──────────────────────────
        self._cell_bar = ctk.CTkFrame(self, height=28, fg_color=("gray85", "gray20"))
        self._cell_bar.grid(row=2, column=0, columnspan=2, sticky="ew", padx=4, pady=(1, 1))
        self._cell_bar.grid_columnconfigure(1, weight=1)
        self._cell_bar.grid_remove()

        self._cell_bar_col_lbl = ctk.CTkLabel(
            self._cell_bar, text="",
            font=ctk.CTkFont(size=11),
            text_color=("gray40", "gray65"),
            anchor="w",
        )
        self._cell_bar_col_lbl.grid(row=0, column=0, padx=(6, 8), sticky="w")

        self._cell_bar_val_lbl = ctk.CTkLabel(
            self._cell_bar, text="",
            font=ctk.CTkFont(size=11),
            text_color=("gray15", "gray85"),
            anchor="w",
        )
        self._cell_bar_val_lbl.grid(row=0, column=1, padx=(0, 4), sticky="ew")

        self._cell_bar_copy_btn = ctk.CTkButton(
            self._cell_bar, text="📋", width=28, height=22,
            fg_color="transparent",
            hover_color=("gray75", "gray35"),
            text_color=("gray10", "gray90"),
            font=ctk.CTkFont(size=12),
            command=self._copy_selected_cell,
        )
        self._cell_bar_copy_btn.grid(row=0, column=2, padx=(0, 4))

        # ── cell overlay canvas (cell border highlight) ───────────────────────
        self._cell_overlay = tk.Canvas(self, highlightthickness=0, bd=0, cursor="")
        for seq in ("<Button-1>", "<ButtonRelease-1>",
                    "<Button-3>", "<ButtonRelease-3>",
                    "<Motion>", "<Leave>", "<MouseWheel>"):
            self._cell_overlay.bind(seq, self._forward_to_tree)

        self._empty_lbl = ctk.CTkLabel(
            self,
            text="(нет данных)",
            font=ctk.CTkFont(size=13),
            text_color=("#808080", theme_colors.accent()),
        )
        self._render()

    def _on_vsb_set(self, first, last):
        if float(first) <= 0.0 and float(last) >= 1.0:
            self._vsb.grid_remove()
        else:
            self._vsb.grid(row=0, column=1, sticky="ns")
        self._vsb.set(first, last)

    def _on_hsb_set(self, first, last):
        if float(first) <= 0.0 and float(last) >= 1.0:
            self._hsb.grid_remove()
        else:
            self._hsb.grid(row=1, column=0, sticky="ew")
        self._hsb.set(first, last)

    # ── стиль (обновляется при смене темы) ───────────────────────────────────

    @staticmethod
    def _apply_style():
        mode   = ctk.get_appearance_mode()
        accent = theme_colors.accent()
        if (mode   == ResultTable._last_applied_mode
                and accent == ResultTable._last_applied_accent):
            return
        ResultTable._last_applied_mode   = mode
        ResultTable._last_applied_accent = accent
        s = ttk.Style()
        try:
            if s.theme_use() in ('vista', 'xpnative', 'winnative'):
                s.theme_use('clam')
        except Exception:
            pass
        dark = mode == "Dark"
        bg   = "#2b2b2b" if dark else "#dbdbdb"
        fg   = "#dcddde" if dark else "#1a1a1a"
        hdr  = "#3a3a3a" if dark else "#c5c5c5"
        sel  = theme_colors.accent()
        s.configure("SS.Treeview",
                    background=bg, foreground=fg, fieldbackground=bg,
                    rowheight=24, borderwidth=0, font=("Segoe UI", 10))
        s.configure("SS.Treeview.Heading",
                    background=hdr, foreground=fg, relief="flat",
                    font=("Segoe UI", 10, "bold"), padding=(6, 3))
        s.map("SS.Treeview",
              background=[("selected", sel)],
              foreground=[("selected", "#ffffff")])
        s.map("SS.Treeview.Heading",
              background=[("active", sel)])

    def refresh_style(self):
        ResultTable._last_applied_mode = ""
        ResultTable._apply_style()
        self._render()

    def update_accent(self, new_accent: str):
        """Обновляет акцентный цвет при живой смене темы."""
        try:
            self._empty_lbl.configure(text_color=("#808080", new_accent))
        except Exception:
            pass
        ResultTable._last_applied_mode = ""
        ResultTable._apply_style()
        self._render()

    # ── данные ───────────────────────────────────────────────────────────────

    def set_data(self, rows: list, columns: list, reset_hidden: bool = True):
        all_rows = [list(r) for r in rows]
        self._columns = list(columns)
        self._sort_col = None
        self._sort_rev = False
        self._current_page = 0
        if reset_hidden:
            self._hidden_keys = set()
            self._hidden_rows = {}
        if self._hidden_keys:
            visible = []
            for r in all_rows:
                key = str(r[0]) if r else ""
                if key in self._hidden_keys:
                    self._hidden_rows.setdefault(key, []).append(r)
                else:
                    visible.append(r)
            self._rows = visible
        else:
            self._rows = all_rows
        self._render()

    def _total_pages(self) -> int:
        if not self._rows:
            return 1
        return max(1, (len(self._rows) + _PAGE_SIZE - 1) // _PAGE_SIZE)

    def _page_rows(self) -> list:
        start = self._current_page * _PAGE_SIZE
        return self._rows[start:start + _PAGE_SIZE]

    def _page_prev(self):
        if self._current_page > 0:
            self._current_page -= 1
            self._render()

    def _page_next(self):
        if self._current_page < self._total_pages() - 1:
            self._current_page += 1
            self._render()

    def _page_first(self):
        if self._current_page != 0:
            self._current_page = 0
            self._render()

    def _page_last(self):
        last = self._total_pages() - 1
        if self._current_page != last:
            self._current_page = last
            self._render()

    def _render(self):
        # Reset cell selection whenever table content changes (page, sort, new data)
        if hasattr(self, "_scroll_redraw_id"):
            try:
                self.after_cancel(self._scroll_redraw_id)
            except Exception:
                pass
        self._clear_cell_border()
        if self._cell_bar.winfo_ismapped():
            self._cell_bar.grid_remove()
        self._focused_item = None
        self._focused_col = None

        self._tree.delete(*self._tree.get_children())
        if not self._columns:
            self._tree.grid_remove()
            self._vsb.grid_remove()
            self._hsb.grid_remove()
            self._pag_bar.grid_remove()
            self._empty_lbl.grid(row=0, column=0, columnspan=2, sticky="nsew")
            return

        ResultTable._apply_style()
        self._empty_lbl.grid_remove()
        self._tree.grid(row=0, column=0, sticky="nsew")

        total = self._total_pages()
        if total > 1:
            self.grid_rowconfigure(3, weight=0)
            self._pag_bar.grid()
            cur = self._current_page
            start_row = cur * _PAGE_SIZE + 1
            end_row   = min((cur + 1) * _PAGE_SIZE, len(self._rows))
            self._pag_label.configure(
                text=f"стр. {cur + 1} / {total}  ({start_row}–{end_row} из {len(self._rows)})"
            )
            self._btn_prev.configure(
                state="normal" if cur > 0 else "disabled",
                text_color=("gray10", "gray90") if cur > 0 else ("gray60", "gray45"),
            )
            self._btn_next.configure(
                state="normal" if cur < total - 1 else "disabled",
                text_color=("gray10", "gray90") if cur < total - 1 else ("gray60", "gray45"),
            )
        else:
            self._pag_bar.grid_remove()

        col_ids = ["_bulb_", "_copy_", "_row_"] + [f"c{i}" for i in range(len(self._columns))]
        self._tree["columns"] = col_ids

        self._tree.heading("_bulb_", text="")
        self._tree.column("_bulb_", width=24, minwidth=24, stretch=False)

        self._tree.heading("_copy_", text="⎘")
        self._tree.column("_copy_", width=26, minwidth=26, stretch=False)

        self._tree.heading("_row_", text="№")
        self._tree.column("_row_", width=40, minwidth=30, stretch=False, anchor="e")

        page_rows = self._page_rows()
        for i, (cid, name) in enumerate(zip(col_ids[3:], self._columns)):
            arrow  = (" ▲" if not self._sort_rev else " ▼") if self._sort_col == i else ""
            self._tree.heading(cid, text=name + arrow,
                               command=lambda c=i: self._sort_by(c))
            sample = [str(r[i]) for r in page_rows[:80] if i < len(r)]
            chars  = max([len(name)] + [len(v) for v in sample], default=4)
            self._tree.column(cid, width=min(max(chars * 8 + 16, 60), 360),
                              minwidth=50, stretch=True)

        dark = ctk.get_appearance_mode() == "Dark"
        self._tree.tag_configure("r0", background="#2b2b2b" if dark else "#dbdbdb")
        self._tree.tag_configure("r1", background="#333333" if dark else "#d0d0d0")
        row_offset = self._current_page * _PAGE_SIZE
        for i, row in enumerate(page_rows):
            vals = ["", "⎘", str(row_offset + i + 1)] + ["NULL" if v is None else str(v) for v in row]
            self._tree.insert("", "end", values=vals, tags=(f"r{i % 2}",))

    # ── сортировка ────────────────────────────────────────────────────────────

    def _sort_by(self, col: int):
        self._sort_rev = (not self._sort_rev) if self._sort_col == col else False
        self._sort_col = col
        self._apply_sort()

    def _apply_sort(self):
        col = self._sort_col
        if col is None:
            self._render()
            return

        def key(row):
            v = row[col] if col < len(row) else None
            if v is None:
                return (2, "")
            try:
                return (0, float(str(v)))
            except (ValueError, TypeError):
                return (1, str(v).lower())

        self._rows.sort(key=key, reverse=self._sort_rev)
        self._current_page = 0
        self._render()

    # ── копирование ───────────────────────────────────────────────────────────

    def _cell_value(self, event) -> Optional[str]:
        item = self._tree.identify_row(event.y)
        col  = self._tree.identify_column(event.x)
        if not item or not col:
            return None
        idx  = int(col.lstrip("#")) - 1
        if idx <= 2:  # _bulb_, _copy_, _row_ columns
            return None
        vals = self._tree.item(item, "values")
        return vals[idx] if idx < len(vals) else None

    def _on_right_click(self, event):
        item = self._tree.identify_row(event.y)
        if not item:
            return
        self._tree.selection_set(item)
        cell_val = self._cell_value(event)
        all_vals = self._tree.item(item, "values")
        row_text = "\t".join(str(v) for v in all_vals[3:])  # skip _bulb_, _copy_, _row_
        menu = tk.Menu(self, tearoff=0)
        if cell_val is not None:
            menu.add_command(label="Копировать ячейку",
                             command=lambda: self._clip(cell_val))
        menu.add_command(label="Копировать строку",
                         command=lambda: self._clip(row_text))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _on_copy_click(self, event):
        col = self._tree.identify_column(event.x)
        if col == "#1":  # _bulb_ column
            item = self._tree.identify_row(event.y)
            if item:
                self._show_bulb_menu(event, item)
        elif col == "#2":  # _copy_ column
            item = self._tree.identify_row(event.y)
            if not item:
                return
            self._copy_row_column_format(item)

    def _copy_row_column_format(self, item):
        vals = self._tree.item(item, "values")
        actual_vals = vals[3:]  # skip _bulb_, _copy_, _row_ columns
        self._clip("\n".join(str(v) for v in actual_vals))
        try:
            self._tree.set(item, "_copy_", "✓")
            self.after(600, lambda: self._restore_copy_icon(item))
        except Exception:
            pass

    def _restore_copy_icon(self, item):
        try:
            if self._tree.exists(item):
                self._tree.set(item, "_copy_", "⎘")
        except Exception:
            pass

    def _on_cell_click(self, event):
        if self._tree.identify_region(event.x, event.y) != "cell":
            return
        col  = self._tree.identify_column(event.x)
        item = self._tree.identify_row(event.y)
        self._focused_col  = col
        self._focused_item = item
        if col in ("#1", "#2", "#3") or not item:
            return
        col_idx = int(col.lstrip("#")) - 1
        vals = self._tree.item(item, "values")
        if col_idx >= len(vals):
            return
        name_idx = col_idx - 3
        col_name = self._columns[name_idx] if 0 <= name_idx < len(self._columns) else col
        cell_val = str(vals[col_idx])
        self._selected_cell_value = cell_val
        self._show_cell_bar(col_name, cell_val)
        self._draw_cell_border(item, col)

    def _on_ctrl_keypress(self, event):
        if event.keycode == 67:
            return self._copy_row(event)

    def _copy_row(self, event=None):
        sel = self._tree.selection()
        if not sel:
            return "break"
        values = self._tree.item(sel[0], "values")
        if self._focused_col and self._focused_col not in ("#1", "#2", "#3"):
            col_idx = int(self._focused_col.lstrip("#")) - 1
            if 0 <= col_idx < len(values):
                self._clip(str(values[col_idx]))
                return "break"
        self._clip("\t".join(str(v) for v in values[3:]))  # skip _bulb_, _copy_, _row_
        return "break"

    # ── cell info bar ─────────────────────────────────────────────────────────

    def _show_cell_bar(self, col_name: str, value: str):
        preview = value if len(value) <= 80 else value[:77] + "…"
        self._cell_bar_col_lbl.configure(text=f"Колонка: {col_name}")
        self._cell_bar_val_lbl.configure(text=preview)
        self._cell_bar_copy_btn.configure(text="📋")
        if not self._cell_bar.winfo_ismapped():
            self._cell_bar.grid()

    def _copy_selected_cell(self):
        if self._selected_cell_value:
            self._clip(self._selected_cell_value)
            self._cell_bar_copy_btn.configure(text="✓")
            self.after(600, lambda: self._cell_bar_copy_btn.configure(text="📋"))

    # ── cell border overlay ───────────────────────────────────────────────────

    def _draw_cell_border(self, item: str, col_id: str):
        self._cell_overlay.place_forget()
        self._cell_overlay.delete("all")
        try:
            bbox = self._tree.bbox(item, col_id)
        except Exception:
            return
        if not bbox:
            return
        x, y, w, h = bbox
        tags = self._tree.item(item, "tags")
        dark = ctk.get_appearance_mode() == "Dark"
        bg = ("#333333" if dark else "#d0d0d0") if "r1" in tags else ("#2b2b2b" if dark else "#dbdbdb")
        fg = "#dcddde" if dark else "#1a1a1a"
        tree_x = self._tree.winfo_x()
        tree_y = self._tree.winfo_y()
        self._cell_overlay.configure(bg=bg)
        self._cell_overlay.place(x=tree_x + x, y=tree_y + y, width=w, height=h)
        self._cell_overlay.lift()
        col_idx = int(col_id.lstrip("#")) - 1
        vals = self._tree.item(item, "values")
        if col_idx < len(vals):
            self._cell_overlay.create_text(
                4, h // 2, text=str(vals[col_idx]),
                fill=fg, font=("Segoe UI", 10), anchor="w",
            )
        self._cell_overlay.create_rectangle(
            1, 1, w - 2, h - 2,
            outline=theme_colors.accent(), width=2, fill="",
        )

    def _clear_cell_border(self):
        self._cell_overlay.delete("all")
        self._cell_overlay.place_forget()

    def _forward_to_tree(self, event):
        dx = self._cell_overlay.winfo_x() - self._tree.winfo_x()
        dy = self._cell_overlay.winfo_y() - self._tree.winfo_y()
        tx, ty = event.x + dx, event.y + dy
        t = event.type
        if t == tk.EventType.ButtonPress:
            self._tree.event_generate(f"<Button-{event.num}>", x=tx, y=ty)
        elif t == tk.EventType.ButtonRelease:
            self._tree.event_generate(f"<ButtonRelease-{event.num}>", x=tx, y=ty)
        elif t == tk.EventType.Motion:
            self._tree.event_generate("<Motion>", x=tx, y=ty)
        elif t == tk.EventType.Leave:
            self._tree.event_generate("<Leave>", x=tx, y=ty)
        elif t == tk.EventType.MouseWheel:
            self._tree.event_generate("<MouseWheel>", x=tx, y=ty, delta=event.delta)

    def _on_scroll_redraw(self, event=None):
        if self._focused_item and self._focused_col:
            if hasattr(self, "_scroll_redraw_id"):
                try:
                    self.after_cancel(self._scroll_redraw_id)
                except Exception:
                    pass
            item, col = self._focused_item, self._focused_col
            self._scroll_redraw_id = self.after(10, lambda: self._draw_cell_border(item, col))

    # ── hover: лампочка ──────────────────────────────────────────────────────

    def _on_motion(self, event):
        item = self._tree.identify_row(event.y)
        if item:
            if item != self._hovered_item:
                self._clear_bulb(self._hovered_item)
                self._hovered_item = item
                try:
                    self._tree.set(item, "_bulb_", "💡")
                except Exception:
                    pass
        else:
            self._clear_bulb(self._hovered_item)
            self._hovered_item = ""

    def _on_leave_tree(self, event):
        self._clear_bulb(self._hovered_item)
        self._hovered_item = ""

    def _clear_bulb(self, item: str):
        if item:
            try:
                if self._tree.exists(item):
                    self._tree.set(item, "_bulb_", "")
            except Exception:
                pass

    # ── контекстное меню лампочки ─────────────────────────────────────────────

    def _show_bulb_menu(self, event, item: str):
        vals = self._tree.item(item, "values")
        first_col_val = str(vals[3]) if len(vals) > 3 else ""
        row_text = "\t".join(str(v) for v in vals[3:])
        menu = tk.Menu(self, tearoff=0)

        top = self.winfo_toplevel()
        if hasattr(top, "_open_reminder_for_row"):
            menu.add_command(
                label="💡 Напомнить",
                command=lambda: top._open_reminder_for_row(first_col_val),
            )
        menu.add_command(label="📋 Копировать стр",
                         command=lambda: self._clip(row_text))
        menu.add_command(label="👁 Следить", state="disabled")
        menu.add_separator()
        menu.add_command(label="🙈 Скрыть",
                         command=lambda: self._hide_row(first_col_val))
        if self._hidden_rows:
            show_menu = tk.Menu(menu, tearoff=0)
            for key in list(self._hidden_rows.keys()):
                show_menu.add_command(label=str(key),
                                      command=lambda k=key: self._show_row(k))
            menu.add_cascade(label="📂 Показать ▶", menu=show_menu)
        else:
            menu.add_command(label="📂 Показать", state="disabled")
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _hide_row(self, first_col_val: str):
        self._hidden_keys.add(first_col_val)
        new_rows = []
        for r in self._rows:
            if r and str(r[0]) == first_col_val:
                self._hidden_rows.setdefault(first_col_val, []).append(r)
            else:
                new_rows.append(r)
        self._rows = new_rows
        self._hovered_item = ""
        self._render()

    def _show_row(self, key: str):
        self._hidden_keys.discard(key)
        if key in self._hidden_rows:
            self._rows.extend(self._hidden_rows.pop(key))
            if self._sort_col is not None:
                self._apply_sort()
                return
        self._render()

    def _clip(self, text: str):
        top = self.winfo_toplevel()
        top.clipboard_clear()
        top.clipboard_append(str(text))

    def export_to_csv(self, filepath: str):
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(self._columns)
            for row in self._rows:
                writer.writerow(["" if v is None else v for v in row])

    def export_to_excel(self, filepath: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            tk.messagebox.showerror(
                "Ошибка экспорта",
                "Библиотека openpyxl не установлена.\n"
                "Выполните: pip install openpyxl",
            )
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты"
        ws.append(list(self._columns))
        header_fill = PatternFill("solid", fgColor="0D9488")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in self._rows:
            ws.append(["" if v is None else v for v in row])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col[:201]), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
        ws.freeze_panes = "A2"
        wb.save(filepath)
